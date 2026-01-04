from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from tabulate import tabulate
import os

FONT_SIZE = 12
FONT_NAME = "Times New Roman"

def formatting_excel(file_path):
    """
    Format seluruh sheet Excel:
    - Font Times New Roman
    - All borders (thin)
    - Auto column width
    """

    if not os.path.exists(file_path):
        print(f"Error: File {file_path} tidak ditemukan.")
        return

    try:
        wb = load_workbook(file_path)

        default_font = Font(name=FONT_NAME, size=FONT_SIZE)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for ws in wb.worksheets:

            for row in ws.iter_rows():
                for cell in row:
                    cell.font = default_font
                    cell.border = thin_border

                    if cell.column == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter

                for cell in col:
                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                ws.column_dimensions[col_letter].width = max_length * 1.2 + 2

        wb.save(file_path)
        print(f"File disimpan di: {file_path}")

    except Exception as e:
        print(f"Terjadi kesalahan saat memproses Excel: {e}")

def preview_table(df, title):
    """
    Menampilkan pratinjau DataFrame di terminal dengan format yang rapi.
    """
    print(f"\n" + "="*50)
    print(f" PREVIEW: {title.upper()} ")
    print("="*50)
    print(tabulate(df, headers='keys', tablefmt='grid', showindex=True))
    print("="*50 + "\n")